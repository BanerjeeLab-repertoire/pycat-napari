"""Per-operation scientific guidance for the Navigator (Method-Widget Spec 3 — the content model).

The Navigator can *choose* an op; Spec 3 lets it *explain* one — when to use it, its advantages and limitations,
the alternatives, when it does not apply, and references. That content is **authored by the scientist**, never
generated: it encodes judgement a machine must not fabricate (when Cellpose beats thresholding is a scientific
call, not a lookup). So this module is deliberately the same shape as ``sections.py``: explicit DATA
(``data/operation_guidance.json``, schema-versioned) + a refuse-to-guess reader. The store ships EMPTY; entries
appear only as the scientist authors them.

Authoring vehicle: :func:`generate_guidance_workbook` writes a fill-in ``.xlsx`` (one row per op, the factual
columns pre-filled, the judgement columns blank); :func:`ingest_guidance_workbook` reads a filled workbook back
into the JSON. The runtime reader (:func:`guidance_for`) needs only ``json``; the workbook helpers import
``openpyxl`` lazily so the hot path stays light.
"""
from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path
from typing import Optional

_GUIDANCE_PATH = Path(__file__).parent / "data" / "operation_guidance.json"

#: The authored fields. ``when_to_use`` is prose; the rest are lists. Kept here so the reader, the workbook
#: generator, and the coverage ratchet agree on one vocabulary.
GUIDANCE_TEXT_FIELDS = ("when_to_use",)
GUIDANCE_LIST_FIELDS = ("advantages", "limitations", "alternatives", "not_applicable_when", "references")
GUIDANCE_FIELDS = GUIDANCE_TEXT_FIELDS + GUIDANCE_LIST_FIELDS


@lru_cache(maxsize=1)
def _guidance() -> dict:
    """The ``guidance`` table from operation_guidance.json (cached). Returns ``{}`` if the file is missing or
    malformed rather than raising at import — an absent store degrades to 'no guidance authored yet', which the
    UI renders as 'not documented', not a crash."""
    try:
        with open(_GUIDANCE_PATH, encoding="utf-8") as fh:
            data = json.load(fh)
    except (OSError, ValueError):
        return {}
    guidance = data.get("guidance", {})
    return guidance if isinstance(guidance, dict) else {}


def guidance_for(op_id: str) -> Optional[dict]:
    """The authored guidance for an op-id, or ``None`` when it has not been authored yet. ``None`` is a real
    answer (no guidance exists), never a fabricated one — the content is authored by hand, never guessed. The
    caller shows 'not documented yet', the same refuse-to-guess discipline as the execution adapters and the
    section bindings."""
    entry = _guidance().get(op_id)
    return dict(entry) if isinstance(entry, dict) else None


def authored_op_ids() -> frozenset:
    """Every op-id that has authored guidance — the coverage ratchet (test_guidance_coverage) reads this."""
    return frozenset(_guidance().keys())


def section_guidance(op_id: str, *, alternatives=None) -> dict:
    """Assemble the pop-out content for one generated-panel section (Method-Widget Spec 4): the op's OWN authored
    guidance plus the same for each alternative it could be swapped for — so the pop-out shows this operation and
    its rivals side by side with their tradeoffs, in place. Pure over the guidance store; the Qt pop-out only
    renders this dict, so the decision of WHAT to show is testable apart from the widget.

    ``alternatives`` — the candidate op-ids to compare against (the caller passes the planner's considered
    candidates, e.g. from :meth:`Planner.explain_segmentation_choice`, so the pop-out lists exactly what the
    planner weighed). When ``None``, falls back to the op's own authored ``alternatives`` field. An unauthored op
    (or alternative) comes back with ``documented=False`` and ``guidance=None`` — the pop-out shows
    'not documented yet', never a fabricated stand-in.

    Returns ``{"op_id", "documented", "guidance", "alternatives": [{"op_id", "documented", "guidance"}, ...]}``."""
    own = guidance_for(op_id)
    alts = list(alternatives) if alternatives is not None else list((own or {}).get("alternatives", []) or [])
    seen = set()
    alt_entries = []
    for a in alts:
        if a == op_id or a in seen:
            continue
        seen.add(a)
        g = guidance_for(a)
        alt_entries.append({"op_id": a, "documented": g is not None, "guidance": g})
    return {"op_id": op_id, "documented": own is not None, "guidance": own, "alternatives": alt_entries}


# ── authoring vehicle: a fill-in workbook, and its ingest back to JSON ────────────────────────────────

def _catalog_rows():
    """(op_id, module, summary) for every catalog op — the factual columns the workbook pre-fills."""
    cat = json.loads((Path(__file__).parent / "data" / "operation_catalog.json").read_text(encoding="utf-8"))
    return [(o.get("op"), o.get("module", ""), o.get("summary", "")) for o in cat.get("operations", [])]


def generate_guidance_workbook(path):
    """Write a fill-in authoring workbook to ``path``: one row per catalog op, with the FACTUAL columns
    (op_id / module / summary) pre-filled and the JUDGEMENT columns (the GUIDANCE_FIELDS) blank for the scientist
    to complete. List fields take one item per line. Returns the row count. Requires ``openpyxl``."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "operation_guidance"
    header = ["op_id", "module", "summary"] + list(GUIDANCE_FIELDS)
    ws.append(header)
    rows = _catalog_rows()
    for op_id, module, summary in rows:
        ws.append([op_id, module, summary] + [""] * len(GUIDANCE_FIELDS))
    wb.save(str(path))
    return len(rows)


def _split_list_cell(value) -> list:
    """A list-field cell → a clean list: one item per line, trimmed, blanks dropped."""
    if value is None:
        return []
    return [part.strip() for part in str(value).replace("\r", "\n").split("\n") if part.strip()]


def ingest_guidance_workbook(path, *, out_path=None) -> dict:
    """Read a filled authoring workbook back into the guidance store and write it to ``out_path`` (default: the
    shipped operation_guidance.json). Only rows with SOME authored content are recorded — a blank row is not a
    guidance entry. Returns the ``{op_id: guidance}`` map written. Requires ``openpyxl``."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active
    header = None
    guidance = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        cells = dict(zip(header, row))
        op_id = cells.get("op_id")
        if not op_id:
            continue
        entry = {}
        for f in GUIDANCE_TEXT_FIELDS:
            text = cells.get(f)
            if text and str(text).strip():
                entry[f] = str(text).strip()
        for f in GUIDANCE_LIST_FIELDS:
            items = _split_list_cell(cells.get(f))
            if items:
                entry[f] = items
        if entry:                                          # a blank row is not a guidance entry
            guidance[str(op_id)] = entry

    target = Path(out_path) if out_path is not None else _GUIDANCE_PATH
    payload = {"schema_version": 1, "guidance": guidance}
    target.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    _guidance.cache_clear()
    return guidance
